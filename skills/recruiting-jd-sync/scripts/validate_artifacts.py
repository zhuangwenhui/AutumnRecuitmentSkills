#!/usr/bin/env python3
"""Validate recruiting JD Markdown, classification Markdown, and workbook IDs."""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(f"openpyxl is required: {exc}")


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def parse_summary_ids(text: str, id_pattern: str) -> set[str]:
    rx = re.compile(rf"^### ({id_pattern})\b", re.M)
    return set(rx.findall(text))


def parse_classification_ids(text: str, id_pattern: str) -> set[str]:
    rx = re.compile(rf"\*\*({id_pattern})\b")
    return set(rx.findall(text))


def parse_workbook_ids(path: Path, id_pattern: str, exclude_sheets: set[str]) -> tuple[list[str], dict[str, list[tuple[str, int]]]]:
    rx = re.compile(rf"^{id_pattern}$")
    wb = load_workbook(path, read_only=True, data_only=True)
    ids: list[str] = []
    locations: dict[str, list[tuple[str, int]]] = {}
    for ws in wb.worksheets:
        if ws.title in exclude_sheets:
            continue
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row, 3).value
            if value is None or value == "":
                continue
            value = str(value)
            if rx.match(value):
                ids.append(value)
                locations.setdefault(value, []).append((ws.title, row))
    return ids, locations


def count_total_marker(text: str) -> int | None:
    match = re.search(r"岗位总计\*\*：\s*(\d+)\s*个", text)
    return int(match.group(1)) if match else None


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--summary", default="岗位JD汇总.md")
    parser.add_argument("--classification", default="岗位JD_按技术栈深度分类.md")
    parser.add_argument("--workbook", default="投递优先级清单.xlsx")
    parser.add_argument("--id-pattern", default=r"D\d+-\d{2}")
    parser.add_argument("--exclude-sheet", action="append", default=["评分说明"])
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    summary_path = Path(args.summary)
    classification_path = Path(args.classification)
    workbook_path = Path(args.workbook)

    summary_text = read_text(summary_path)
    classification_text = read_text(classification_path)
    summary_ids = parse_summary_ids(summary_text, args.id_pattern)
    classification_ids = parse_classification_ids(classification_text, args.id_pattern)
    workbook_ids, locations = parse_workbook_ids(workbook_path, args.id_pattern, set(args.exclude_sheet))
    workbook_id_set = set(workbook_ids)
    duplicates = sorted([k for k, v in Counter(workbook_ids).items() if v > 1])

    issues: list[str] = []
    if summary_ids != classification_ids:
        issues.append("summary/classification ID sets differ")
    if summary_ids != workbook_id_set:
        issues.append("summary/workbook ID sets differ")
    if duplicates:
        issues.append("duplicate workbook IDs")
    for label, text in [("summary", summary_text), ("classification", classification_text)]:
        if "原编号" in text:
            issues.append(f"{label} contains 原编号")
        if "BOSS兴趣" in text:
            issues.append(f"{label} contains BOSS兴趣")
    for label, text, ids in [
        ("summary", summary_text, summary_ids),
        ("classification", classification_text, classification_ids),
    ]:
        total = count_total_marker(text)
        if total is not None and total != len(ids):
            issues.append(f"{label} total marker {total} != ID count {len(ids)}")

    result = {
        "ok": not issues,
        "issues": issues,
        "counts": {
            "summary": len(summary_ids),
            "classification": len(classification_ids),
            "workbook_rows": len(workbook_ids),
            "workbook_unique": len(workbook_id_set),
        },
        "diffs": {
            "summary_minus_classification": sorted(summary_ids - classification_ids),
            "classification_minus_summary": sorted(classification_ids - summary_ids),
            "summary_minus_workbook": sorted(summary_ids - workbook_id_set),
            "workbook_minus_summary": sorted(workbook_id_set - summary_ids),
            "workbook_duplicates": duplicates,
        },
        "locations": locations,
    }

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print("OK" if result["ok"] else "FAILED")
        print(json.dumps(result["counts"], ensure_ascii=False, indent=2))
        for issue in issues:
            print(f"- {issue}")
        for key, values in result["diffs"].items():
            if values:
                print(f"{key}: {values}")
    return 0 if result["ok"] else 1


if __name__ == "__main__":
    sys.exit(main())
