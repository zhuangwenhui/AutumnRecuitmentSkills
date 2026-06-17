#!/usr/bin/env python3
"""Scan a recruiting workbook for sheet counts, sources, links, fills, and style drift."""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


GREEN_RGBS = {"FF92D050", "FF00B050", "FFCCFFCC", "FFC6EFCE"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", nargs="?", default="投递优先级清单.xlsx")
    parser.add_argument("--exclude-sheet", action="append", default=["评分说明"])
    parser.add_argument("--new-id", action="append", default=[], help="Inventory ID to check for submitted-fill leakage.")
    parser.add_argument("--company-keyword", action="append", default=[], help="Company/group keyword expected to be isolated.")
    parser.add_argument("--json", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    wb = load_workbook(Path(args.workbook), data_only=False)
    url_re = re.compile(r"^https?://", re.I)
    id_re = re.compile(r"^D\d+-\d{2}$")
    new_ids = set(args.new_id)

    ids: list[str] = []
    sheet_counts: dict[str, int] = {}
    source_counts: Counter[str] = Counter()
    keyword_hits: dict[str, list[tuple[str, int, str, str, str]]] = {k: [] for k in args.company_keyword}
    nonweb_hyperlinks = []
    blank_row_links = []
    bad_link_style = []
    bad_filters = []
    bad_new_fills = []

    for ws in wb.worksheets:
        if ws.title in args.exclude_sheet:
            continue
        sheet_counts[ws.title] = max(ws.max_row - 1, 0)
        expected_filter = f"A1:O{ws.max_row}"
        if ws.auto_filter.ref and ws.auto_filter.ref != expected_filter:
            bad_filters.append((ws.title, ws.auto_filter.ref, expected_filter))
        for row in range(2, ws.max_row + 1):
            values = [ws.cell(row, col).value for col in range(1, 16)]
            did = str(ws.cell(row, 3).value or "")
            company = str(ws.cell(row, 4).value or "")
            title = str(ws.cell(row, 5).value or "")
            source = ws.cell(row, 11).value
            if did:
                ids.append(did)
            if source is not None:
                source_counts[str(source)] += 1
            if not any(v not in (None, "") for v in values[:11] + values[12:]):
                link_cell = ws.cell(row, 12)
                if link_cell.value or link_cell.hyperlink:
                    blank_row_links.append((ws.title, row, link_cell.value, link_cell.hyperlink.target if link_cell.hyperlink else None))
            for keyword in args.company_keyword:
                if keyword in company or keyword in title:
                    keyword_hits[keyword].append((ws.title, row, did, company, title))
            for col in range(1, 16):
                cell = ws.cell(row, col)
                if cell.hyperlink:
                    value = str(cell.value or "")
                    target = str(cell.hyperlink.target or "")
                    if not (url_re.match(value) or url_re.match(target)):
                        nonweb_hyperlinks.append((ws.title, cell.coordinate, value, target))
                    if cell.font.name != "微软雅黑" or float(cell.font.sz or 0) != 10.0 or cell.alignment.horizontal != "center":
                        bad_link_style.append((ws.title, cell.coordinate, value, cell.font.name, cell.font.sz, cell.alignment.horizontal))
            if did in new_ids:
                for col in range(1, 16):
                    rgb = getattr(ws.cell(row, col).fill.fgColor, "rgb", None)
                    if rgb in GREEN_RGBS:
                        bad_new_fills.append((ws.title, row, ws.cell(row, col).coordinate, rgb))

    duplicates = sorted([k for k, v in Counter([x for x in ids if id_re.match(x)]).items() if v > 1])
    result = {
        "sheet_counts": sheet_counts,
        "total_rows": sum(sheet_counts.values()),
        "id_rows": len(ids),
        "unique_ids": len(set(ids)),
        "duplicate_ids": duplicates,
        "source_counts": dict(source_counts),
        "keyword_hits": keyword_hits,
        "nonweb_hyperlinks": nonweb_hyperlinks,
        "blank_row_links": blank_row_links,
        "bad_link_style": bad_link_style,
        "bad_filters": bad_filters,
        "bad_new_fills": bad_new_fills,
    }
    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    return 1 if any(result[k] for k in ["duplicate_ids", "nonweb_hyperlinks", "blank_row_links", "bad_filters", "bad_new_fills"]) else 0


if __name__ == "__main__":
    sys.exit(main())
