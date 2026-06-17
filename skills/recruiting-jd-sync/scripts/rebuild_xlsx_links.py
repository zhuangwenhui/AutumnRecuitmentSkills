#!/usr/bin/env python3
"""Rebuild workbook link-column hyperlinks from the JD dictionary by inventory ID."""

from __future__ import annotations

import argparse
import re
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def parse_urls(summary_path: Path, id_pattern: str) -> dict[str, str]:
    text = summary_path.read_text(encoding="utf-8")
    urls: dict[str, str] = {}
    for section in re.split(r"^### ", text, flags=re.M)[1:]:
        first = section.splitlines()[0] if section.splitlines() else ""
        m = re.match(rf"({id_pattern})\b", first)
        if not m:
            continue
        did = m.group(1)
        um = re.search(r"^\| URL \| (.*?) \|$", section, flags=re.M)
        if not um:
            continue
        field = um.group(1).strip()
        lm = re.search(r"\((https?://[^)]+)\)", field)
        if lm:
            urls[did] = lm.group(1)
        elif field.startswith(("http://", "https://")):
            urls[did] = field
    return urls


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--summary", default="岗位JD汇总.md")
    parser.add_argument("--workbook", default="投递优先级清单.xlsx")
    parser.add_argument("--id-pattern", default=r"D\d+-\d{2}")
    parser.add_argument("--exclude-sheet", action="append", default=["评分说明"])
    parser.add_argument("--display", default="查看JD")
    args = parser.parse_args()

    urls = parse_urls(Path(args.summary), args.id_pattern)
    wb = load_workbook(args.workbook)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    normal_font = Font(name="微软雅黑", size=10, color="000000")
    hyperlink_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
    rebuilt = 0
    missing = []

    for ws in wb.worksheets:
        if ws.title in set(args.exclude_sheet):
            continue
        for row in range(2, ws.max_row + 1):
            did = str(ws.cell(row, 3).value or "")
            if not did:
                continue
            cell = ws.cell(row, 12)
            cell.hyperlink = None
            target = urls.get(did)
            if target:
                if not cell.value or str(cell.value).startswith(("http://", "https://")):
                    cell.value = args.display
                cell.hyperlink = target
                cell.font = copy(hyperlink_font)
                rebuilt += 1
            else:
                cell.font = copy(normal_font)
                if str(cell.value or "").startswith(("http://", "https://")):
                    missing.append((ws.title, row, did, cell.value))
            cell.alignment = alignment

    wb.save(args.workbook)
    print(f"rebuilt={rebuilt}")
    if missing:
        print("missing_url_for_ids:")
        for item in missing:
            print(item)
    return 0 if not missing else 1


if __name__ == "__main__":
    sys.exit(main())
