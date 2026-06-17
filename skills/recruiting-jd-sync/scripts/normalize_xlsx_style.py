#!/usr/bin/env python3
"""Normalize recruiting workbook font, alignment, filters, and optional orphan rows."""

from __future__ import annotations

import argparse
import re
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", nargs="?", default="投递优先级清单.xlsx")
    parser.add_argument("--exclude-sheet", action="append", default=["评分说明"])
    parser.add_argument("--font", default="微软雅黑")
    parser.add_argument("--size", type=float, default=10.0)
    parser.add_argument("--cols", type=int, default=15)
    parser.add_argument("--remove-orphan-link-rows", action="store_true")
    args = parser.parse_args()

    wb = load_workbook(Path(args.workbook))
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    normal_font = Font(name=args.font, size=args.size, color="000000")
    hyperlink_font = Font(name=args.font, size=args.size, color="0563C1", underline="single")
    header_font = Font(name=args.font, size=args.size, bold=True, color="FFFFFF")
    id_re = re.compile(r"^D(\d+)-(\d{2})$")
    removed = []

    for ws in wb.worksheets:
        if ws.title in set(args.exclude_sheet):
            continue
        if args.remove_orphan_link_rows:
            for row in range(ws.max_row, 1, -1):
                values = [ws.cell(row, col).value for col in range(1, args.cols + 1)]
                nonempty = [v for v in values if v not in (None, "")]
                if len(nonempty) == 0 or (len(nonempty) == 1 and ws.cell(row, 12).value not in (None, "")):
                    removed.append((ws.title, row, ws.cell(row, 12).value))
                    ws.delete_rows(row, 1)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:O{ws.max_row}"
        for col in range(1, args.cols + 1):
            cell = ws.cell(1, col)
            cell.font = copy(header_font)
            cell.alignment = alignment
        for row in range(2, ws.max_row + 1):
            did = str(ws.cell(row, 3).value or "")
            for col in range(1, args.cols + 1):
                cell = ws.cell(row, col)
                cell.alignment = alignment
                cell.font = copy(hyperlink_font if cell.hyperlink else normal_font)
                if col == 15:
                    match = id_re.match(did)
                    if match:
                        cell.value = int(match.group(1)) + int(match.group(2)) / 100
                    cell.number_format = "0.00"

    wb.save(args.workbook)
    print(f"normalized={args.workbook}")
    if removed:
        print("removed_orphan_rows:")
        for item in removed:
            print(item)
    return 0


if __name__ == "__main__":
    sys.exit(main())
